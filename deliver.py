from email.policy import default
from http.client import NOT_FOUND
from operator import not_
import os
from turtle import end_fill
from numpy import true_divide
import openpyxl
from datetime import timedelta
from routing import routing
from path_generation import cpg
import numpy as np
import pickle
import copy

path_not_get = 0

def process_data(dataname):
    dir_path = os.path.dirname(os.path.realpath(__file__))
    path_to_excel = dir_path + "\\"+dataname
    excel = openpyxl.load_workbook(path_to_excel,data_only=True,keep_vba=False)
    sheet = excel.active
    head = 'A'
    tail = '1'
    data_type = []
    data = []
    max_column = int(sheet.max_column)
    for i in range(max_column):
        index = head + tail
        value = sheet[index].value
        data_type.append(value)
        head = chr(ord(head)+1)
    max_row = int(sheet.max_row) - 1
    for i in range(max_row):
        head = 'A'
        tail = str(i+2)
        data.append({})
        row = data[i]
        for j in range(max_column):
              type_name = data_type[j]
              index = head + tail
              value = sheet[index].value
              head = chr(ord(head)+1) 
              row[type_name] = value
    return data       
      

    
def time_convert(time):
    time = time.split('~')
    left = time[0].split(':')
    left_hour = left[0]
    left_hour = int(left_hour)
    left_min = left[1]
    left_min = int(left_min)
    right = time[1].split(':')
    right_hour = right[0]
    right_min = right[1]
    right_hour = int(right_hour)
    right_min = int(right_min)
    time_arr =[]
    time_left = timedelta(hours=left_hour,minutes=left_min)
    time_right = timedelta(hours=right_hour,minutes=right_min)
    time_arr.append(time_left)
    time_arr.append(time_right)
    return time_arr

def convert_sec_time(data):
    time = timedelta(seconds = data)
    return time

def find_duration(path_group,start,end):
    for path in path_group:
        if path['start']==start and path['end'] == end:
            return path['duration']
    return -1

def create_timeslot(start_time,end_time):
    timeslot = []
    tmp_start_time = start_time
    tmp_end_time = start_time + timedelta(minutes = 30)
    while tmp_end_time <= end_time:
        time_interval = {'start':tmp_start_time,'end':tmp_end_time}
        timeslot.append(time_interval)
        tmp_start_time += timedelta(minutes = 30)
        tmp_end_time += timedelta(minutes = 30)

    return timeslot

def get_order_time_interval(order):
    time = time_convert(order['時間段'])
    start = time[0]
    end = time[1]
    time_interval = {'start':start,'end':end}
    return time_interval

def ini_drivers(driver_num):
    drivers = []
    for i in range(driver_num):
        driver = {'order':[],'path':[]}
        drivers.append(driver)

    return drivers

def free_timeslot_for_order(driver_job,order):
    index = -1
    judge = True
    l = len(driver_job['order'])
    order_time_interval = get_order_time_interval(order)
    for i in range(l):
        if driver_job['order'][i]['time_interval'] == order_time_interval:
            judge = False
            break
        elif driver_job['order'][i]['time_interval']['end'] > order_time_interval['end']:
            index = i-1
            break
    return judge,index


def continuous_timeslot(time1,time2):
    time1_end = time1['end']
    time2_start = time2['start']
    if time1_end == time2_start:
        return 0  # 0 indicate the two timeslots are continuous
    elif time2_start > time1_end:
        return 1   # 1 indicate the new timeslot is much further
    else:
        return -1 #indicate the new timeslot is the same

def ini_array(num):
    a = []
    for i in range(num):
        a.append(i)
    return a

def generate_order_cluster_group(order_group):  #create possible combination of the order cluster group
    l = len(order_group)
    load_index = ini_array(l)
    order_cluster_group = []
    while load_index:
        index = load_index[0]
        order_cluster_l = 1
        order_cluster =[]
        order_cluster.append(order_group[index])
        load_index.remove(index)
        p = 0
        while p < len(load_index) and order_cluster_l <3:
            last_order = order_cluster[-1]
            last_order_timeslot = get_order_time_interval(last_order)
            index = load_index[p]
            order = order_group[index]
            order_timeslot = get_order_time_interval(order)
            judge = continuous_timeslot(last_order_timeslot,order_timeslot)
            if judge == 0:
                order_cluster.append(order)
                load_index.remove(index)
                order_cluster_l += 1
            elif judge == 1:
                break
            p += 1
        order_cluster_group.append(order_cluster)    
    return order_cluster_group
    

def create_path_cluster(order_cluster):
    try:       #try to read the path record
       file = open('path.pkl','rb')
       path_record = pickle.load(file)
       file.close()
    except:   #if no path record generation_path_grouppcgthen create one
       path_record = []
    order_length = len(order_cluster)  #the length of the orders
    location_group = [] #create the group that contain location sequences of the path
    for i in range(order_length):   #
        start = order_cluster[i]['哪裡取']
        location_group.append(start)
    for i in range(order_length):
        end = order_cluster[i]['送到']
        location_group.append(end)
    location_group_length = int(order_length*2) #generate location path group length
    path_cluster_group = cpg.generation_path_group(location_group_length) 
    pcg_length = len(path_cluster_group)
    pcg_weight = []
    
    for i in range(pcg_length):  #load the total path within paths group
       path = path_cluster_group[i]  
       path_length = len(path)
       start_i = path[0] # way start index
       weight = timedelta()
       
       for i in range(1,path_length):
          end_i = path[i]  # way end index
          start = location_group[start_i] #way start location
          end = location_group[end_i] #way end location
          judge = find_duration(path_record,start,end)
          if judge != -1:  #way in path record
             weight += judge 
          else:   #way not in path record
             global path_not_get
             path_not_get += 1
             new_path = routing.create_path(start,end)
             path_record.append(new_path) 
             weight += new_path['duration']
             
          start_i = end_i #now the way end become new way start

       pcg_weight.append(weight)
      
    #find the minimum time cost in total path weight matrix
    path_combination_index = np.argmin(pcg_weight) #the index of the mimimum value in the weight matrix
    #find the min cost list in the value matrix
    path_combination = path_cluster_group[path_combination_index] 
    start_i = path_combination[0]
    start = location_group[start_i]
    time_interval = get_order_time_interval(order_cluster[-1])
    start_time = time_interval['start']
    end_time = time_interval['end']
    time_interval = {'start':start_time,'end':end_time}
    path_store = {'time_interval':time_interval,'main':[]} # the path record
    ini_time_cost = timedelta()
    path = {'start':start,'end':start,'duration':ini_time_cost,'status':'take','order':order_cluster[start_i]}
    path_store['main'].append(path)
    for i in range(1,location_group_length):
        end_i = path_combination[i]
        start = location_group[start_i]
        end = location_group[end_i]
        
        if end_i < order_length:
            status = 'take'
        else:
            status = 'deliver'
        if status == 'take':
           order_index = end_i
        elif status =='deliver':
           order_index = end_i - order_length
        order = order_cluster[order_index]
        if status == 'take':
            default_time_cost = timedelta()
        else:
            if order['配送上門'] == '是':
               default_time_cost = timedelta(minutes = 10)
            else:
               default_time_cost = timedelta(minutes =5)
        
        path = {'start':start,'end':end,'duration':find_duration(path_record,start,end)+default_time_cost,'status':status,'order':order_cluster[order_index]}
        path_store['main'].append(path)
        start_i = end_i

    file = open('path.pkl','wb')
    pickle.dump(path_record,file)
    file.close()
    return path_store

def create_path_group(order_cluster_group):
    path_group = []
    for order_cluster in order_cluster_group:
        path_cluster = create_path_cluster(order_cluster)
        path_group.append(path_cluster)
    return path_group

def allocate_job(path_group,drivers_num):
    drivers_job = []
    for i in range(drivers_num):
         drivers_job.append([])
    path_group_copy = copy.deepcopy(path_group)
    for path_cluster in path_group:
        path_cluster_main = path_cluster['main']
        for i in range (drivers_num):
            driver_job = drivers_job[i]
            if len(driver_job) == 0:
                driver_job.append(path_cluster)
                try:
                   path_group_copy.remove(path_cluster)
                   break
                except:
                   pass
                
            else:
                last_path_cluster = driver_job[-1]
                last_order = last_path_cluster['main'][-1]['order']
                last_order_timeslot = get_order_time_interval(last_order)
                this_order = path_cluster['main'][-1]['order']
                this_order_timeslot = get_order_time_interval(this_order)
                judge = continuous_timeslot(last_order_timeslot,this_order_timeslot)
                if judge == 1:
                    driver_job.append(path_cluster)
                    path_group_copy.remove(path_cluster)
                    break
    
    return drivers_job,path_group_copy

def check_path_cluster_deliver_status(path_cluster):
    time_comsumpt = timedelta()
    not_on_time_count = 0
    path_cluster_main = path_cluster['main']
    time_limit = timedelta(minutes = 15)
    for path in path_cluster_main:
        time_comsumpt += path['duration']
        if path['status'] == 'deliver':
            time_limit += timedelta(minutes = 30)    
            if time_comsumpt > time_limit:
                global max_time_delay
                max_time_delay = max(max_time_delay,time_comsumpt - time_limit)
                not_on_time_count += 1
    return time_comsumpt, not_on_time_count  
            
def check_path_group_deliver_status(path_group):
    time_comsumpt = timedelta()
    not_on_time_count = 0
    for path_cluster in path_group:
        new_time_comsumpt,new_not_on_time_count = check_path_cluster_deliver_status(path_cluster)
        time_comsumpt += new_time_comsumpt
        not_on_time_count += new_not_on_time_count
    
    return time_comsumpt,not_on_time_count

def seperate_day_data(unprocess_order_group):
    process_order_group = []
    daily_order_group = []
    for order in unprocess_order_group:  #search each order
        delivery_date = order['配送日期']   #take current order delivery date
        try:        #do if there be previous order
            if delivery_date == old_delivery_date:
                daily_order_group.append(order)
            else:     #new delivery date
                process_order_group.append(daily_order_group)
                delivery_order_group = []
                old_delivery_date = delivery_date
        except:
            old_delivery_date = delivery_date
            daily_order_group.append(order)
    if daily_order_group:
        process_order_group.append(daily_order_group)

def check_order_cluster_group_deliver_status(process_order_group):
    time_comsumpt = 0
    not_no_time_count = 0
    for order_cluster in process_order_group:
        get_time_comsumpt,get_not_on_time_count = check_path_group_deliver(stat)

def process_order_group(order_group):


def main():
    unprocess_order_group = process_data('data_full.xlsx')
    process_order_group = generate_order_cluster_group(unprocess_order_group)
    

main()